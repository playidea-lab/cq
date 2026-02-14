"""XLSX Parser - openpyxl 기반 + ZIP 이미지 추출."""

import logging
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from c4.c2.parsers.base import BaseParser, ImageData, ParseResult
from c4.c2.parsers.ir_models import (
    CellStyle,
    Document,
    MergeInfo,
    create_heading,
    create_image,
    create_table,
)
from c4.c2.parsers.utils.chart_parser import parse_chart_xml
from c4.c2.parsers.utils.image import generate_image_id, get_mime_from_extension

logger = logging.getLogger(__name__)


class XlsxParser(BaseParser):
    """XLSX 문서 파서.

    파싱 규칙:
    - 모든 시트 처리 (시트명을 heading으로)
    - 첫 행 header
    - 나머지 body
    - 머지셀 지원
    """

    @property
    def supported_extensions(self) -> list[str]:
        return [".xlsx"]  # .xls는 xls_parser에서 처리

    def parse(self, file_path: Path) -> Document:
        """XLSX 파일을 IR로 변환."""
        result = self.parse_with_images(file_path)
        return result.document

    def parse_with_images(self, file_path: Path) -> ParseResult:
        """XLSX 파일을 IR과 이미지로 변환.

        XLSX는 ZIP 구조이므로 xl/media/ 폴더에서 이미지 직접 추출.
        """
        # read_only=False로 변경하여 머지셀 정보 접근
        wb = load_workbook(str(file_path), read_only=False, data_only=True)
        blocks = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 여러 시트가 있으면 시트명을 heading으로
            if len(wb.sheetnames) > 1:
                blocks.append(create_heading(2, sheet_name))

            table_block = self._parse_sheet(ws)
            if table_block:
                blocks.append(table_block)

        wb.close()

        # ZIP에서 이미지 추출
        extracted_images, image_blocks = self._extract_images_from_zip(file_path)

        # 차트 추출 (ZIP 구조에서)
        chart_blocks, chart_images = self._extract_charts(file_path, len(extracted_images))
        extracted_images.extend(chart_images)

        # 이미지 블록을 문서 끝에 추가
        blocks.extend(image_blocks)

        # 차트 블록 추가 (테이블 + 이미지)
        blocks.extend(chart_blocks)

        return ParseResult(document=Document(blocks=blocks), images=extracted_images)

    def _extract_images_from_zip(self, file_path: Path) -> tuple[list[ImageData], list]:
        """ZIP 구조에서 이미지 파일 추출.

        XLSX 내부 구조:
        - xl/media/image1.png
        - xl/media/image2.jpeg
        - ...

        Returns:
            (ImageData 목록, Image 블록 목록)
        """
        extracted_images: list[ImageData] = []
        image_blocks = []

        try:
            with zipfile.ZipFile(file_path, "r") as zf:
                # xl/media/ 폴더의 파일들 찾기
                media_files = [
                    name for name in zf.namelist()
                    if name.startswith("xl/media/") and not name.endswith("/")
                ]

                for idx, media_path in enumerate(sorted(media_files)):
                    try:
                        # 파일 데이터 읽기
                        image_data = zf.read(media_path)

                        # 확장자에서 MIME 타입 결정
                        ext = Path(media_path).suffix.lstrip(".")
                        mime_type, normalized_ext = get_mime_from_extension(ext)

                        # 지원하지 않는 형식 스킵
                        if normalized_ext == ".bin":
                            continue

                        # 너무 작은 파일 스킵 (아이콘 등)
                        if len(image_data) < 1000:
                            continue

                        # 이미지 ID 생성
                        image_id = generate_image_id(image_data, idx)
                        image_id_with_ext = f"{image_id}{normalized_ext}"

                        # IR 블록 및 데이터 생성
                        image_blocks.append(
                            create_image(image_id=image_id_with_ext, mime_type=mime_type)
                        )
                        extracted_images.append(
                            ImageData(
                                image_id=image_id_with_ext,
                                data=image_data,
                                mime_type=mime_type,
                            )
                        )

                    except Exception as e:
                        # 개별 이미지 추출 실패 시 스킵
                        logger.debug("Failed to extract media image %s: %s", media_path, e)
                        continue

        except zipfile.BadZipFile:
            # ZIP이 아니거나 손상된 경우 이미지 없이 반환
            pass

        return extracted_images, image_blocks

    def _parse_sheet(self, ws: Worksheet):
        """시트를 테이블 블록으로 변환."""
        # 머지셀 정보 추출
        merge_info = self._extract_merge_info(ws)

        # 데이터 범위 확인
        if ws.max_row is None or ws.max_row == 0:
            return None

        rows_data = []
        rows_styles: list[list[CellStyle]] = []

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            row_data = []
            row_style = []
            for cell in row:
                # 값 추출
                value = cell.value
                if value is None:
                    row_data.append("")
                else:
                    row_data.append(str(value))

                # 스타일 추출
                cell_style = self._extract_cell_style(cell)
                row_style.append(cell_style)

            rows_data.append(row_data)
            rows_styles.append(row_style)

        if not rows_data:
            return None

        # 빈 행 제거 (모든 셀이 빈 경우) - 스타일도 함께 제거
        filtered_data = []
        filtered_styles = []
        for row_data, row_style in zip(rows_data, rows_styles):
            if any(cell.strip() for cell in row_data):
                filtered_data.append(row_data)
                filtered_styles.append(row_style)

        if not filtered_data:
            return None

        rows_data = filtered_data
        rows_styles = filtered_styles

        # 첫 행은 header
        header = rows_data[0]
        body_rows = rows_data[1:] if len(rows_data) > 1 else []

        return create_table(
            header=header,
            rows=body_rows,
            merge_info=merge_info if merge_info else None,
            cell_styles=rows_styles if rows_styles else None,
        )

    def _extract_cell_style(self, cell) -> CellStyle:
        """셀에서 스타일 정보 추출."""
        background_color = None
        is_bold = False
        font_size = None
        text_align = None

        try:
            # 배경색 추출
            fill = cell.fill
            if fill and fill.fgColor and fill.fgColor.rgb:
                rgb = fill.fgColor.rgb
                # ARGB 형식 (8자리) 또는 RGB (6자리)
                if isinstance(rgb, str) and len(rgb) >= 6:
                    # "00000000" (투명) 또는 "FFFFFFFF" (흰색)이 아닌 경우
                    if rgb not in ("00000000", "FFFFFFFF", "00FFFFFF"):
                        # ARGB면 앞 2자리 제거
                        color = rgb[-6:] if len(rgb) == 8 else rgb
                        background_color = f"#{color}"

            # 볼드 추출
            font = cell.font
            if font and font.bold:
                is_bold = True

            # 폰트 크기 추출
            if font and font.size:
                font_size = float(font.size)

            # 정렬 추출
            alignment = cell.alignment
            if alignment and alignment.horizontal:
                text_align = alignment.horizontal

        except Exception as e:
            # 스타일 추출 실패 시 기본값 사용
            logger.debug("Failed to extract cell style: %s", e)

        return CellStyle(
            background_color=background_color,
            is_bold=is_bold,
            font_size=font_size,
            text_align=text_align,
        )

    def _extract_merge_info(self, ws: Worksheet) -> list[MergeInfo]:
        """시트에서 머지셀 정보 추출."""
        merge_info = []

        for merged_range in ws.merged_cells.ranges:
            # merged_range는 CellRange 객체
            min_row = merged_range.min_row - 1  # 0-indexed로 변환
            min_col = merged_range.min_col - 1
            rowspan = merged_range.max_row - merged_range.min_row + 1
            colspan = merged_range.max_col - merged_range.min_col + 1

            if rowspan > 1 or colspan > 1:
                merge_info.append(
                    MergeInfo(
                        row=min_row,
                        col=min_col,
                        rowspan=rowspan,
                        colspan=colspan,
                    )
                )

        return merge_info

    def _extract_charts(
        self, file_path: Path, start_index: int
    ) -> tuple[list, list[ImageData]]:
        """XLSX ZIP 구조에서 차트 추출.

        XLSX 차트 구조:
        - xl/charts/chart1.xml (차트 데이터)
        - xl/media/image*.png (차트 이미지)

        Returns:
            (차트 블록 목록, 이미지 데이터 목록)
        """
        blocks = []
        images = []

        try:
            with zipfile.ZipFile(file_path, "r") as zf:
                # 차트 XML 파일 찾기
                chart_files = sorted(
                    [f for f in zf.namelist() if f.startswith("xl/charts/chart") and f.endswith(".xml")]
                )

                for chart_file in chart_files:
                    try:
                        chart_content = zf.read(chart_file).decode("utf-8")
                        chart_block = parse_chart_xml(chart_content)
                        if chart_block:
                            blocks.append(chart_block)
                    except Exception as e:
                        logger.debug("Failed to parse chart %s: %s", chart_file, e)

                # 차트 이미지는 이미 _extract_images_from_zip에서 추출되므로
                # 여기서는 차트 데이터 테이블만 추출

        except zipfile.BadZipFile:
            pass

        return blocks, images
